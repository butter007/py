# -*- coding: utf-8 -*- 
import xlrd
import datetime
import openpyxl
from openpyxl.styles import Font,colors,Color

# 签收库里的数据导入
txt = open("D:\py\sign.txt")
li = []
for line in txt:
    if line.startswith('LW') or line.startswith('SUZJ') or line.startswith("K25"):
        # 去掉签收库导出的数据的换行符
        line = line[:-1]
 #       line = line.encode('utf-8')
        li.append(line)
txt.close()

# print li[1]
# print li
# target = open("D:\py\out.txt",'a+');
# for str in li :
#    target.write(str)
# target.close()

# 部门做账数据列表
listthreelw = []
listthreesuzj = []
listsevenlw = []
listsevensuzj = []
listtenlw = []
listtensuzj = []

# 帐里需要做销售的待运
data = xlrd.open_workbook("D:\py\daiyun.xlsx")
table = data.sheets()[0]
nrows = table.nrows
col = table.col_values(1)
# print col

# 账里需要做销售的待运在签收库里面有的待运
temp = []
# 账里需要做销售的待运未签收的待运
unsignedtemp = []
for data in col:
    if (li.count(data) > 0):
        temp.append(data)
    else:
        unsignedtemp.append(data)

# print li.count('SUZJ17070076')




for data in temp:
    # strid = table.row_values(j)[1]
    if data.startswith('LW1703') or data.startswith('LW1803'):
        listthreelw.append(data)
    elif data.startswith('SUZJ1703') or data.startswith('SUZJ1803'):
        listthreesuzj.append(data)
    elif data.startswith('LW1707') or data.startswith('LW1807'):
        listsevenlw.append(data)
    elif data.startswith('SUZJ1707') or data.startswith('SUZJ1807'):
        listsevensuzj.append(data)
    elif data.startswith('LW1710') or data.startswith('LW1810'):
        listtenlw.append(data)
    elif data.startswith('SUZJ1710') or data.startswith('SUZJ1810'):
        listtensuzj.append(data)

excle_Workbook = openpyxl.Workbook()
excel_sheet_name = datetime.datetime.now().strftime('%Y-%m-%d')
excel_sheet = excle_Workbook.active
excel_sheet.title = excel_sheet_name

index = 1
if len(listthreelw) != 0:
    for data in listthreelw:
        excel_sheet.cell(index, 2, data)
        index += 1

if len(listthreesuzj) != 0:
    for data in listthreesuzj:
        excel_sheet.cell(index, 2, data)
        index += 1

excel_sheet.cell(index, 2, '')
index+=1
if len(listsevenlw) != 0:
    for data in listsevenlw:
        excel_sheet.cell(index, 2, data)
        index += 1

if len(listsevensuzj) != 0:
    for data in listsevensuzj:
        excel_sheet.cell(index, 2, data)
        index += 1
excel_sheet.cell(index, 2, '')
index+=1
if len(listtenlw) != 0:
    for data in listtenlw:
        excel_sheet.cell(index, 2, data)
        index += 1

if len(listtensuzj) != 0:
    for data in listtensuzj:
        excel_sheet.cell(index, 2, data)
        index += 1

unsignindex=1
if unsignedtemp:
    for data in unsignedtemp:
    	if data.startswith('LW1703') or data.startswith('LW1803')or data.startswith('SUZJ1703') or data.startswith('SUZJ1803') or data.startswith('LW1707') or data.startswith('LW1807') or data.startswith('SUZJ1707') or data.startswith('SUZJ1807') or data.startswith('LW1710') or data.startswith('LW1810') or data.startswith('SUZJ1710') or data.startswith('SUZJ1810'):
        	excel_sheet.cell(unsignindex, 6, data)
        	unsignindex += 1


excle_Workbook.save('D:\py\pest.xlsx')
