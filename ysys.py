# -*- coding: utf-8 -*- 
import xlrd
import datetime
import openpyxl
from openpyxl.styles import Font,colors,Color

# 导入预收外汇账款
yushoudata = xlrd.open_workbook("D:\py\yushouzk.xlsx")
yushoutable = yushoudata.sheets()[0]
yushounrows = yushoutable.nrows
yushoucol = yushoutable.col_values(1)

yingshoudata = xlrd.open_workbook("D:\py\yingshouzk.xlsx")
yinghsoutable = yingshoudata.sheets()[0]
yingshounrows = yinghsoutable.nrows
yingshoucol = yinghsoutable.col_values(1)

alikelist = []
for i in range(1, yushounrows):
    for j in range(1, yingshounrows):
        if yushoutable.row_values(i)[1] == yinghsoutable.row_values(j)[1]:
            alikelist.append(yushoutable.row_values(i)[1:16])
            li = yinghsoutable.row_values(j)
            del (li[2])
            alikelist.append(li[1:16])
            # print yushoutable.row_values(i)[13]
            # print li[12]
            m = yushoutable.row_values(i)[13] - li[13]
            alikelist.append(['', '', '', '', '', '', '', '', '', '', '', '', m, '', ''])

print(len(alikelist))

excle_Workbook = openpyxl.Workbook()
excel_sheet_name = datetime.datetime.now().strftime('%Y-%m-%d')
excel_sheet = excle_Workbook.active
excel_sheet.title = excel_sheet_name

index = 1
# pattern = Pattern()
# # 设置底纹颜色
# pattern.pattern_back_colour = 0x0A
# print (type(pattern))
# style0 = XFStyle()
# style0.pattern = pattern
#
# redfontstyle = XFStyle()
# bluefontstyle = XFStyle()
# pinkfontstyle = XFStyle()
#
# redfontstyle.font.colour_index = 2
# bluefontstyle.font.colour_index = 4
# pinkfontstyle.font.colour_index = 6
# print (type(redfontstyle))
if len(alikelist) != 0:
    for data in alikelist:
        for i in range(len(data)):
            if (index) % 3 == 0 and i == 12:
                if data[i] < -1000 or data[i] > 1000:
                    excel_sheet.cell(index, i + 1, data[i])
                    excel_sheet.cell(index, i + 1).font = Font(color=colors.RED)
                elif -100 < data[i] and data[i] < 100:
                    excel_sheet.cell(index, i + 1, data[i])
                    excel_sheet.cell(index, i + 1).font = Font(color=colors.BLUE)
                elif -1000 <= data[i] < -100 or 100 < data[i] <= 1000:
                    excel_sheet.cell(index, i + 1, data[i])
                    excel_sheet.cell(index, i + 1).font = Font(color='00FF00FF')
            else:
                excel_sheet.cell(index, i + 1, data[i])
        index += 1

excle_Workbook.save('D:\py\ysyswhh.xlsx')
